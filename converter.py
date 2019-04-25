import os
import datetime
import xlrd
from openpyxl import load_workbook

def convert_fond_csv(filename, input_directory):
    # initialization
    INPUT_DIRECTORY = input_directory
    TEMPLATE_DIRECTORY = str(os.path.dirname(os.path.abspath(__file__))) + "\\template\\"
    OUTPUT_DIRECTORY = str(os.path.dirname(os.path.abspath(__file__))) + "\\output\\"

    # input file
    FILE_INPUT_NAME = filename
    FILE_INPUT_LOCATION = INPUT_DIRECTORY + FILE_INPUT_NAME
    wb = xlrd.open_workbook(FILE_INPUT_LOCATION)
    sheet = wb.sheet_by_index(0)

    # Select document type
    document_type = ''
    if 'Portfolio ID KVG' in sheet.cell_value(9, 0):
        if 'CDS Type' in sheet.cell_value(25, 0):
            document_type = 'CDS'
        elif 'Single Currency Interest Rate SWAP' in sheet.cell_value(25, 1):
            document_type = 'SWAP'
    elif 'Transref. AM' in sheet.cell_value(9, 0):
        if 'FUTURE' in sheet.cell_value(10, 11):
            document_type = 'FUTURE'
        else:
            document_type = 'BOND'
    elif 'SPOT/FORWARD' in sheet.cell_value(7, 22):
        document_type = 'DTG'

    print(document_type)

    # data init
    kvg_bic = ''
    kvg_name = ''
    manager_id = ''
    manager_name = ''
    newm_canc = ''
    portfolio_id = ''
    portfolio_id_kvg = ''
    portfolio_name = ''
    buy_sell = ''
    opep_clop = ''
    quantity = ''
    instr_id = ''
    finan_instr_id = ''
    finan_instr_name = ''
    upi = ''
    price = ''
    ccy = ''
    date = ''
    exec_broker_id_type = ''
    exec_broker_id = ''
    exec_broker_name = ''

    # bond
    transfer_am = ''
    sec_id_type = ''
    sec_id = ''
    sec_name = ''
    interest = ''
    settlement_date = ''
    settlement_ccy = ''

    # DTG
    cleared_trade = ''
    counterparty_name = ''
    counterparty_type = ''
    counterparty_id = ''
    buy_ccy = ''
    buy_quantity = ''
    sell_ccy = ''
    sell_quantity = ''
    spot_forward = ''
    exchange_rate = ''
    non_deliverable = ''
    exec_timestamp = ''

    transfer_am_2 = ''
    newm_canc_2 = ''
    portfolio_id_kvg_2 = ''
    portfolio_name_2 = ''
    cleared_trade_2 = ''
    counterparty_name_2 = ''
    counterparty_type_2 = ''
    counterparty_id_2 = ''
    buy_ccy_2 = ''
    buy_quantity_2 = ''
    sell_ccy_2 = ''
    sell_quantity_2 = ''
    spot_forward_2 = ''
    exchange_rate_2 = ''
    date_2 = ''
    settlement_date_2 = ''
    non_deliverable_2 = ''
    exec_timestamp_2 = ''

    transfer_am_3 = ''
    newm_canc_3 = ''
    portfolio_id_kvg_3 = ''
    portfolio_name_3 = ''
    cleared_trade_3 = ''
    counterparty_name_3 = ''
    counterparty_type_3 = ''
    counterparty_id_3 = ''
    buy_ccy_3 = ''
    buy_quantity_3 = ''
    sell_ccy_3 = ''
    sell_quantity_3 = ''
    spot_forward_3 = ''
    exchange_rate_3 = ''
    date_3 = ''
    settlement_date_3= ''
    non_deliverable_3 = ''
    exec_timestamp_3 = ''

    transfer_am_4 = ''
    newm_canc_4 = ''
    portfolio_id_kvg_4 = ''
    portfolio_name_4 = ''
    cleared_trade_4 = ''
    counterparty_name_4 = ''
    counterparty_type_4 = ''
    counterparty_id_4 = ''
    buy_ccy_4 = ''
    buy_quantity_4 = ''
    sell_ccy_4 = ''
    sell_quantity_4 = ''
    spot_forward_4 = ''
    exchange_rate_4 = ''
    date_4 = ''
    settlement_date_4 = ''
    non_deliverable_4 = ''
    exec_timestamp_4 = ''

    # SWAP
    product_type = ''
    op_inc_dec_cl = ''
    effective_date = ''
    termination_date = ''
    business_days = ''

    rate_payer = ''
    notional_amount = ''
    payment_frequency = ''
    first_payment_date = ''
    rate = ''
    day_count_fraction = ''
    calc_period_end_dates_adjustment = ''
    business_days_following = ''

    ccy_2 = ''
    rate_payer_2 = ''
    rate_receiver = ''
    notional_amount_2 = ''
    payment_frequency_2 = ''
    first_payment_date_2 = ''
    calc_period_end_dates_adjustment_2 = ''
    business_days_2 = ''
    reset_frequency = ''
    floating_rate_option = ''
    spread_in_bps = ''
    day_count_fraction_2 = ''
    payment = ''

    # CDS
    reference_entity_name = ''
    reference_obligation_type = ''
    reference_obligation_ID = ''
    scheduled_term_date = ''
    fixed_rate_deal_spread = ''
    settlement_method = ''
    settlement_business_days = ''
    trade_effective_date = ''
    actual_spread = ''
    buy_sell_protection = ''
    payer = ''
    payee = ''
    payment_ccy = ''
    payment_amount = ''
    payment_settlement_date = ''
    first_fee_payment = ''
    accrual_start_date = ''
    new_notional_amount = ''

    # data extraction
    list_input_data = []

    if document_type == 'FUTURE':
        # header
        kvg_bic = sheet.cell_value(1, 1)
        kvg_name = sheet.cell_value(2, 1)
        manager_id = int(sheet.cell_value(4, 1))
        manager_name = sheet.cell_value(5, 1)

        # row = 10
        newm_canc = sheet.cell_value(10, 1)
        portfolio_id = sheet.cell_value(10, 3)
        portfolio_id_kvg = sheet.cell_value(10, 4)
        portfolio_name = sheet.cell_value(10, 5)
        buy_sell = sheet.cell_value(10, 6)
        opep_clop = sheet.cell_value(10, 7)
        quantity = int(sheet.cell_value(10, 8))
        instr_id = sheet.cell_value(10, 9)
        finan_instr_id = sheet.cell_value(10, 10)
        finan_instr_name = sheet.cell_value(10, 11)
        upi = sheet.cell_value(10, 12)
        price = str(float(str(sheet.cell_value(10, 14)).replace(",", "."))).replace(".", ",")
        ccy = sheet.cell_value(10, 19)
        date = sheet.cell_value(10, 22)
        exec_broker_id_type = sheet.cell_value(10, 26)
        exec_broker_id = sheet.cell_value(10, 27)
        exec_broker_name = sheet.cell_value(10, 28)

        list_input_data = [kvg_bic, kvg_name, manager_id, manager_name, buy_sell, opep_clop,
                           quantity, finan_instr_id, finan_instr_name, upi, price, ccy, date,
                           exec_broker_id_type, exec_broker_id, exec_broker_name]
    elif document_type == 'BOND':
        # header
        kvg_bic = sheet.cell_value(1, 1)
        kvg_name = sheet.cell_value(2, 1)
        manager_id = int(sheet.cell_value(4, 1))
        manager_name = sheet.cell_value(5, 1)
        transfer_am = sheet.cell_value(10, 0)

        # row = 10
        newm_canc = sheet.cell_value(10, 1)
        portfolio_id = sheet.cell_value(10, 3)
        portfolio_id_kvg = sheet.cell_value(10, 4)
        portfolio_name = sheet.cell_value(10, 5)
        buy_sell = sheet.cell_value(10, 6)
        quantity = str(float(str(sheet.cell_value(10, 7)).replace(",", "."))).replace(".", ",")
        sec_id_type = sheet.cell_value(10, 8)
        sec_id = sheet.cell_value(10, 9)
        sec_name = sheet.cell_value(10, 10)
        price = str(float(str(sheet.cell_value(10, 11)).replace(",", "."))).replace(".", ",")
        ccy = sheet.cell_value(10, 12)
        interest = sheet.cell_value(10, 17)
        date = sheet.cell_value(10, 20)
        settlement_date = sheet.cell_value(10, 22)
        settlement_ccy = sheet.cell_value(10, 23)
        exec_broker_id_type = sheet.cell_value(10, 25)
        exec_broker_id = sheet.cell_value(10, 26)
        exec_broker_name = sheet.cell_value(10, 27)

        list_input_data = [kvg_bic, kvg_name, manager_id, manager_name, transfer_am, newm_canc,
                           portfolio_id, portfolio_id_kvg, portfolio_name, buy_sell, quantity, sec_id_type, sec_id,
                           sec_name, price, ccy, interest, date, settlement_date, settlement_ccy, exec_broker_id_type,
                           exec_broker_id, exec_broker_name]
    elif document_type == 'DTG':
        # header
        kvg_bic = sheet.cell_value(0, 1)
        kvg_name = sheet.cell_value(1, 1)
        manager_id = int(sheet.cell_value(3, 1))
        manager_name = sheet.cell_value(4, 1)

        row_one = 8
        transfer_am = sheet.cell_value(row_one, 0)
        newm_canc = sheet.cell_value(row_one, 1)
        portfolio_id_kvg = sheet.cell_value(row_one, 4)
        portfolio_name = sheet.cell_value(row_one, 5)
        cleared_trade = sheet.cell_value(row_one, 6)
        counterparty_name = sheet.cell_value(row_one, 13)
        counterparty_type = sheet.cell_value(row_one, 14)
        counterparty_id = sheet.cell_value(row_one, 15)
        buy_ccy = sheet.cell_value(row_one, 18)
        buy_quantity = str(float(str(sheet.cell_value(row_one, 19)).replace(",", "."))).replace(".", ",")
        sell_ccy = sheet.cell_value(row_one, 20)
        sell_quantity = str(float(str(sheet.cell_value(row_one, 21)).replace(",", "."))).replace(".", ",")
        spot_forward = sheet.cell_value(row_one, 22)
        exchange_rate = str(float(str(sheet.cell_value(row_one, 23)).replace(",", "."))).replace(".", ",")
        date = sheet.cell_value(row_one, 24)
        settlement_date = sheet.cell_value(row_one, 25)
        non_deliverable = sheet.cell_value(row_one, 26)
        exec_timestamp = sheet.cell_value(row_one, 27)

        try:
            row_two = 9
            transfer_am_2 = sheet.cell_value(row_two, 0)
            newm_canc_2 = sheet.cell_value(row_two, 1)
            portfolio_id_kvg_2 = sheet.cell_value(row_two, 4)
            portfolio_name_2 = sheet.cell_value(row_two, 5)
            cleared_trade_2 = sheet.cell_value(row_two, 6)
            counterparty_name_2 = sheet.cell_value(row_two, 13)
            counterparty_type_2 = sheet.cell_value(row_two, 14)
            counterparty_id_2 = sheet.cell_value(row_two, 15)
            buy_ccy_2 = sheet.cell_value(row_two, 18)
            buy_quantity_2 = str(float(str(sheet.cell_value(row_two, 19)).replace(",", "."))).replace(".", ",")
            sell_ccy_2 = sheet.cell_value(row_two, 20)
            sell_quantity_2 = str(float(str(sheet.cell_value(row_two, 21)).replace(",", "."))).replace(".", ",")
            spot_forward_2 = sheet.cell_value(row_two, 22)
            exchange_rate_2 = str(float(str(sheet.cell_value(row_two, 23)).replace(",", "."))).replace(".", ",")
            date_2 = sheet.cell_value(row_two, 24)
            settlement_date_2 = sheet.cell_value(row_two, 25)
            non_deliverable_2 = sheet.cell_value(row_two, 26)
            exec_timestamp_2 = sheet.cell_value(row_two, 27)
        except:
            print('could not read row two')

        try:
            row_three = 10
            transfer_am_3 = sheet.cell_value(row_three, 0)
            newm_canc_3 = sheet.cell_value(row_three, 1)
            portfolio_id_kvg_3 = sheet.cell_value(row_three, 4)
            portfolio_name_3 = sheet.cell_value(row_three, 5)
            cleared_trade_3 = sheet.cell_value(row_three, 6)
            counterparty_name_3 = sheet.cell_value(row_three, 13)
            counterparty_type_3 = sheet.cell_value(row_three, 14)
            counterparty_id_3 = sheet.cell_value(row_three, 15)
            buy_ccy_3 = sheet.cell_value(row_three, 18)
            buy_quantity_3 = str(float(str(sheet.cell_value(row_three, 19)).replace(",", "."))).replace(".", ",")
            sell_ccy_3 = sheet.cell_value(row_three, 20)
            sell_quantity_3 = str(float(str(sheet.cell_value(row_three, 21)).replace(",", "."))).replace(".", ",")
            spot_forward_3 = sheet.cell_value(row_three, 22)
            exchange_rate_3 = str(float(str(sheet.cell_value(row_three, 23)).replace(",", "."))).replace(".", ",")
            date_3 = sheet.cell_value(row_three, 24)
            settlement_date_3 = sheet.cell_value(row_three, 25)
            non_deliverable_3 = sheet.cell_value(row_three, 26)
            exec_timestamp_3 = sheet.cell_value(row_three, 27)
        except:
            print('could not read row three')

        try:
            row_four = 11
            transfer_am_4 = sheet.cell_value(row_four, 0)
            newm_canc_4 = sheet.cell_value(row_four, 1)
            portfolio_id_kvg_4 = sheet.cell_value(row_four, 4)
            portfolio_name_4 = sheet.cell_value(row_four, 5)
            cleared_trade_4 = sheet.cell_value(row_four, 6)
            counterparty_name_4 = sheet.cell_value(row_four, 13)
            counterparty_type_4 = sheet.cell_value(row_four, 14)
            counterparty_id_4 = sheet.cell_value(row_four, 15)
            buy_ccy_4 = sheet.cell_value(row_four, 18)
            buy_quantity_4 = str(float(str(sheet.cell_value(row_four, 19)).replace(",", "."))).replace(".", ",")
            sell_ccy_4 = sheet.cell_value(row_four, 20)
            sell_quantity_4 = str(float(str(sheet.cell_value(row_four, 21)).replace(",", "."))).replace(".", ",")
            spot_forward_4 = sheet.cell_value(row_four, 22)
            exchange_rate_4 = str(float(str(sheet.cell_value(row_four, 23)).replace(",", "."))).replace(".", ",")
            date_4 = sheet.cell_value(row_four, 24)
            settlement_date_4 = sheet.cell_value(row_four, 25)
            non_deliverable_4 = sheet.cell_value(row_four, 26)
            exec_timestamp_4 = sheet.cell_value(row_four, 27)
        except:
            print('could not read row four')

        list_input_data = [kvg_bic, kvg_name, manager_id, manager_name, transfer_am, newm_canc,
                           portfolio_id_kvg, portfolio_name, cleared_trade, counterparty_name, counterparty_type,
                           counterparty_id, buy_ccy, buy_quantity, sell_ccy, sell_quantity, spot_forward, exchange_rate,
                           date, settlement_date, non_deliverable, exec_timestamp]
    elif document_type == 'SWAP':
        # header
        kvg_bic = sheet.cell_value(0, 1)
        kvg_name = sheet.cell_value(1, 1)
        manager_id = int(sheet.cell_value(3, 1))
        manager_name = sheet.cell_value(4, 1)

        # row = 8
        column = 1
        portfolio_id = sheet.cell_value(8, column)
        portfolio_id_kvg = sheet.cell_value(9, column)
        portfolio_name = sheet.cell_value(10, column)
        cleared_trade = sheet.cell_value(11, column)
        counterparty_name = sheet.cell_value(18, column)
        counterparty_type = sheet.cell_value(19, column)
        counterparty_id = sheet.cell_value(20, column)
        product_type = sheet.cell_value(25, column)
        transfer_am = str(float(str(sheet.cell_value(30, column)).replace(",", "."))).replace(".", ",")
        newm_canc = sheet.cell_value(32, column)
        op_inc_dec_cl = sheet.cell_value(33, column)
        date = sheet.cell_value(34, column)
        effective_date = sheet.cell_value(36, column)
        termination_date = sheet.cell_value(37, column)
        ccy = sheet.cell_value(42, column)
        rate_payer = sheet.cell_value(43, column)
        notional_amount = str(float(str(sheet.cell_value(44, column)).replace(",", "."))).replace(".", ",")
        payment_frequency = sheet.cell_value(46, column)
        first_payment_date = sheet.cell_value(47, column)
        rate = str(float(str(sheet.cell_value(48, column)).replace(",", "."))).replace(".", ",")
        day_count_fraction = sheet.cell_value(49, column)
        calc_period_end_dates_adjustment = sheet.cell_value(50, column)
        business_days = sheet.cell_value(42, column)
        reset_frequency = sheet.cell_value(63, column)
        floating_rate_option = sheet.cell_value(64, column)
        spread_in_bps = str(float(str(sheet.cell_value(65, column)).replace(",", "."))).replace(".", ",")
        day_count_fraction_2 = sheet.cell_value(67, column)
        payment = sheet.cell_value(68, column)
        rate_payer_2 = sheet.cell_value(70, column)
        rate_receiver = sheet.cell_value(71, column)
        ccy_2 = sheet.cell_value(72, column)
        notional_amount_2 = str(float(str(sheet.cell_value(73, column)).replace(",", "."))).replace(".", ",")
        first_payment_date_2 = sheet.cell_value(74, column)
        exec_timestamp = sheet.cell_value(76, column)

        list_input_data = [portfolio_id, portfolio_id_kvg, portfolio_name, cleared_trade, counterparty_name,
                           counterparty_type, counterparty_id, transfer_am, transfer_am, newm_canc, op_inc_dec_cl, date,
                           effective_date, termination_date, ccy, rate_payer, notional_amount, payment_frequency,
                           first_payment_date, rate, day_count_fraction, calc_period_end_dates_adjustment, business_days,
                           reset_frequency, floating_rate_option, spread_in_bps, day_count_fraction_2, payment,
                           rate_payer_2, rate_receiver, ccy_2, notional_amount_2, first_payment_date_2, exec_timestamp]
    elif document_type == 'CDS':
        # header
        kvg_bic = sheet.cell_value(0, 1)
        kvg_name = sheet.cell_value(1, 1)
        manager_id = int(sheet.cell_value(3, 1))
        manager_name = sheet.cell_value(4, 1)

        # row = 8
        column = 1
        portfolio_id = sheet.cell_value(8, column)
        portfolio_id_kvg = sheet.cell_value(9, column)
        portfolio_name = sheet.cell_value(10, column)
        cleared_trade = sheet.cell_value(11, column)
        counterparty_name = sheet.cell_value(18, column)
        counterparty_type = sheet.cell_value(19, column)
        counterparty_id = sheet.cell_value(20, column)
        product_type = sheet.cell_value(25, column)
        reference_entity_name = sheet.cell_value(30, column)
        reference_obligation_type = sheet.cell_value(31, column)
        reference_obligation_ID = sheet.cell_value(32, column)
        scheduled_term_date = sheet.cell_value(40, column)
        fixed_rate_deal_spread = str(float(str(sheet.cell_value(41, column)).replace(",", "."))).replace(".", ",")
        business_days = sheet.cell_value(42, column)
        payment_frequency = sheet.cell_value(43, column)
        day_count_fraction = sheet.cell_value(44, column)
        settlement_method = sheet.cell_value(45, column)
        settlement_ccy = sheet.cell_value(46, column)
        settlement_business_days = sheet.cell_value(47, column)
        transfer_am = sheet.cell_value(48, column)
        newm_canc = sheet.cell_value(50, column)
        op_inc_dec_cl = sheet.cell_value(51, column)
        date = sheet.cell_value(52, column)
        trade_effective_date = sheet.cell_value(53, column)
        first_fee_payment = sheet.cell_value(54, column)
        accrual_start_date = sheet.cell_value(55, column)
        actual_spread = str(float(str(sheet.cell_value(56, column)).replace(",", "."))).replace(".", ",")
        buy_sell_protection = sheet.cell_value(57, column)
        notional_amount = str(float(str(sheet.cell_value(58, column)).replace(",", "."))).replace(".", ",")
        new_notional_amount = str(float(str(sheet.cell_value(59, column)).replace(",", "."))).replace(".", ",")
        payer = sheet.cell_value(60, column)
        payee = sheet.cell_value(61, column)
        payment_ccy = sheet.cell_value(62, column)
        payment_amount = sheet.cell_value(63, column)
        payment_settlement_date = sheet.cell_value(64, column)
        exec_timestamp = sheet.cell_value(66, column)

        list_input_data = [portfolio_id, portfolio_id_kvg, portfolio_name, cleared_trade, counterparty_name,
                           counterparty_type, counterparty_id, product_type, reference_entity_name,
                           reference_obligation_type, reference_obligation_type, reference_obligation_ID,
                           scheduled_term_date, fixed_rate_deal_spread, business_days, payment_frequency, day_count_fraction,
                           settlement_method, settlement_ccy, settlement_business_days, transfer_am, newm_canc,
                           op_inc_dec_cl, date, trade_effective_date, first_fee_payment, first_fee_payment,
                           accrual_start_date, actual_spread,buy_sell_protection, notional_amount, new_notional_amount,
                           payer, payee, payment_ccy, payment_amount, payment_settlement_date, exec_timestamp]

    else:
        print('NO DOCUMENT TYPE FOUND')

    # template file
    FILE_TEMPLATE_NAME = "template_" + document_type + ".xlsx"
    FILE_TEMPLATE_LOCATION = TEMPLATE_DIRECTORY + FILE_TEMPLATE_NAME
    wb = load_workbook(FILE_TEMPLATE_LOCATION)
    list_worksheets = wb.sheetnames


    # output file
    FILE_OUTPUT_NAME = "output_" + str(document_type) + "_" + str(datetime.datetime.now().time()).replace(":", "_").split(".")[0] + ".xlsx"
    FILE_OUTPUT_LOCATION = OUTPUT_DIRECTORY + FILE_OUTPUT_NAME

    if document_type == 'FUTURE':
        ws = wb["future"]
        ws['A2'] = FILE_INPUT_NAME
        ws['B3'] = kvg_bic
        ws['B4'] = kvg_name
        ws['B6'] = manager_id
        ws['B7'] = manager_name
        ws['A12'] = newm_canc
        ws['B12'] = portfolio_id
        ws['C12'] = portfolio_id_kvg
        ws['D12'] = portfolio_name
        ws['E12'] = buy_sell
        ws['F12'] = opep_clop
        ws['G12'] = quantity
        ws['H12'] = instr_id
        ws['I12'] = finan_instr_id
        ws['J12'] = finan_instr_name
        ws['K12'] = upi
        ws['A16'] = price
        ws['B16'] = ccy
        ws['C16'] = date
        ws['D16'] = exec_broker_id_type
        ws['E16'] = exec_broker_id
        ws['F16'] = exec_broker_name
    elif document_type == 'BOND':
        ws = wb["bond"]
        ws['A2'] = FILE_INPUT_NAME
        ws['B3'] = kvg_bic
        ws['B4'] = kvg_name
        ws['B6'] = manager_id
        ws['B7'] = manager_name
        ws['A12'] = transfer_am
        ws['B12'] = newm_canc
        ws['C12'] = portfolio_id
        ws['D12'] = portfolio_id_kvg
        ws['E12'] = portfolio_name
        ws['F12'] = buy_sell
        ws['G12'] = quantity
        ws['H12'] = sec_id_type
        ws['I12'] = sec_id
        ws['J12'] = sec_name
        ws['A16'] = price
        ws['B16'] = ccy
        ws['C16'] = interest
        ws['D16'] = date
        ws['E16'] = settlement_date
        ws['F16'] = settlement_ccy
        ws['G16'] = exec_broker_id_type
        ws['H16'] = exec_broker_id
        ws['I16'] = exec_broker_name
    elif document_type == 'DTG':
        ws = wb["DTG"]
        ws['A2'] = FILE_INPUT_NAME
        ws['B3'] = kvg_bic
        ws['B4'] = kvg_name
        ws['B6'] = manager_id
        ws['B7'] = manager_name

        row_one_upper = str(12)
        row_one_bottom = str(18)
        ws['A' + row_one_upper] = transfer_am
        ws['B' + row_one_upper] = newm_canc
        ws['D' + row_one_upper] = portfolio_id_kvg
        ws['E' + row_one_upper] = portfolio_name
        ws['F' + row_one_upper] = cleared_trade
        ws['G' + row_one_upper] = counterparty_name
        ws['H' + row_one_upper] = counterparty_type
        ws['I' + row_one_upper] = counterparty_id
        ws['A' + row_one_bottom] = buy_ccy
        ws['B' + row_one_bottom] = buy_quantity
        ws['C' + row_one_bottom] = sell_ccy
        ws['D' + row_one_bottom] = sell_quantity
        ws['E' + row_one_bottom] = spot_forward
        ws['F' + row_one_bottom] = exchange_rate
        ws['G' + row_one_bottom] = date
        ws['H' + row_one_bottom] = settlement_date
        ws['I' + row_one_bottom] = non_deliverable
        ws['J' + row_one_bottom] = exec_timestamp

        row_two_upper = str(13)
        row_two_bottom = str(19)
        ws['A' + row_two_upper] = transfer_am_2
        ws['B' + row_two_upper] = newm_canc_2
        ws['D' + row_two_upper] = portfolio_id_kvg_2
        ws['E' + row_two_upper] = portfolio_name_2
        ws['F' + row_two_upper] = cleared_trade_2
        ws['G' + row_two_upper] = counterparty_name_2
        ws['H' + row_two_upper] = counterparty_type_2
        ws['I' + row_two_upper] = counterparty_id_2
        ws['A' + row_two_bottom] = buy_ccy_2
        ws['B' + row_two_bottom] = buy_quantity_2
        ws['C' + row_two_bottom] = sell_ccy_2
        ws['D' + row_two_bottom] = sell_quantity_2
        ws['E' + row_two_bottom] = spot_forward_2
        ws['F' + row_two_bottom] = exchange_rate_2
        ws['G' + row_two_bottom] = date_2
        ws['H' + row_two_bottom] = settlement_date_2
        ws['I' + row_two_bottom] = non_deliverable_2
        ws['J' + row_two_bottom] = exec_timestamp_2

        row_three_upper = str(14)
        row_three_bottom = str(20)
        ws['A' + row_three_upper] = transfer_am_3
        ws['B' + row_three_upper] = newm_canc_3
        ws['D' + row_three_upper] = portfolio_id_kvg_3
        ws['E' + row_three_upper] = portfolio_name_3
        ws['F' + row_three_upper] = cleared_trade_3
        ws['G' + row_three_upper] = counterparty_name_3
        ws['H' + row_three_upper] = counterparty_type_3
        ws['I' + row_three_upper] = counterparty_id_3
        ws['A' + row_three_bottom] = buy_ccy_3
        ws['B' + row_three_bottom] = buy_quantity_3
        ws['C' + row_three_bottom] = sell_ccy_3
        ws['D' + row_three_bottom] = sell_quantity_3
        ws['E' + row_three_bottom] = spot_forward_3
        ws['F' + row_three_bottom] = exchange_rate_3
        ws['G' + row_three_bottom] = date_3
        ws['H' + row_three_bottom] = settlement_date_3
        ws['I' + row_three_bottom] = non_deliverable_3
        ws['J' + row_three_bottom] = exec_timestamp_3


        row_four_upper = str(15)
        row_four_bottom = str(21)
        ws['A' + row_four_upper] = transfer_am_4
        ws['B' + row_four_upper] = newm_canc_4
        ws['D' + row_four_upper] = portfolio_id_kvg_4
        ws['E' + row_four_upper] = portfolio_name_4
        ws['F' + row_four_upper] = cleared_trade_4
        ws['G' + row_four_upper] = counterparty_name_4
        ws['H' + row_four_upper] = counterparty_type_4
        ws['I' + row_four_upper] = counterparty_id_4
        ws['A' + row_four_bottom] = buy_ccy_4
        ws['B' + row_four_bottom] = buy_quantity_4
        ws['C' + row_four_bottom] = sell_ccy_4
        ws['D' + row_four_bottom] = sell_quantity_4
        ws['E' + row_four_bottom] = spot_forward_4
        ws['F' + row_four_bottom] = exchange_rate_4
        ws['G' + row_four_bottom] = date_4
        ws['H' + row_four_bottom] = settlement_date_4
        ws['I' + row_four_bottom] = non_deliverable_4
        ws['J' + row_four_bottom] = exec_timestamp_4

    elif document_type == 'SWAP':
        ws = wb["SWAP"]
        ws['A1'] = FILE_INPUT_NAME
        ws['B2'] = kvg_bic
        ws['B3'] = kvg_name
        ws['B5'] = manager_id
        ws['B6'] = manager_name
        ws['B10'] = portfolio_id
        ws['B11'] = portfolio_id_kvg
        ws['B12'] = portfolio_name
        ws['B13'] = cleared_trade
        ws['B20'] = counterparty_name
        ws['B21'] = counterparty_type
        ws['B22'] = counterparty_id
        ws['B27'] = product_type
        ws['B32'] = transfer_am
        ws['B34'] = newm_canc
        ws['B35'] = op_inc_dec_cl
        ws['B36'] = date
        ws['B38'] = effective_date
        ws['B39'] = termination_date
        ws['B44'] = ccy_2
        ws['B45'] = rate_payer_2
        ws['B46'] = notional_amount_2
        ws['B48'] = payment_frequency_2
        ws['B49'] = first_payment_date_2
        ws['B50'] = rate
        ws['B51'] = day_count_fraction_2
        ws['B52'] = calc_period_end_dates_adjustment
        ws['B53'] = business_days
        ws['B55'] = ccy_2
        ws['B56'] = rate_payer_2
        ws['B57'] = notional_amount_2
        ws['B59'] = payment_frequency_2
        ws['B60'] = first_payment_date_2
        ws['B61'] = calc_period_end_dates_adjustment_2
        ws['B62'] = business_days_2
        ws['B65'] = reset_frequency
        ws['B66'] = floating_rate_option
        ws['B67'] = spread_in_bps
        ws['B69'] = day_count_fraction_2
        ws['B70'] = payment
        ws['B78'] = exec_timestamp
    elif document_type == 'CDS':
        ws = wb["CDS"]
        ws['A1'] = FILE_INPUT_NAME
        ws['B2'] = kvg_bic
        ws['B3'] = kvg_name
        ws['B5'] = manager_id
        ws['B6'] = manager_name
        ws['B10'] = portfolio_id
        ws['B11'] = portfolio_id_kvg
        ws['B12'] = portfolio_name
        ws['B13'] = cleared_trade
        ws['B20'] = counterparty_name
        ws['B21'] = counterparty_type
        ws['B22'] = counterparty_id
        ws['B27'] = product_type
        ws['B32'] = reference_entity_name
        ws['B33'] = reference_obligation_type
        ws['B34'] = reference_obligation_ID
        ws['B42'] = scheduled_term_date
        ws['B43'] = fixed_rate_deal_spread
        ws['B44'] = business_days
        ws['B45'] = payment_frequency
        ws['B46'] = day_count_fraction
        ws['B47'] = settlement_method
        ws['B48'] = settlement_ccy
        ws['B49'] = settlement_business_days
        ws['B50'] = transfer_am
        ws['B52'] = newm_canc
        ws['B53'] = op_inc_dec_cl
        ws['B54'] = date
        ws['B55'] = trade_effective_date
        ws['B56'] = first_fee_payment
        ws['B57'] = accrual_start_date
        ws['B58'] = actual_spread
        ws['B59'] = buy_sell_protection
        ws['B60'] = notional_amount
        ws['B61'] = new_notional_amount
        ws['B62'] = payer
        ws['B63'] = payee
        ws['B64'] = payment_ccy
        ws['B65'] = payment_amount
        ws['B66'] = payment_settlement_date
        ws['B68'] = exec_timestamp


    print(list_worksheets)
    print(list_input_data)

    wb.save(FILE_OUTPUT_LOCATION)

input_directory = str(os.path.dirname(os.path.abspath(__file__))) + "\\input\\"
for filename in os.listdir(input_directory):
    convert_fond_csv(filename, input_directory)
